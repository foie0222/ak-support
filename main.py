import sys
import time
import openpyxl
from odds import get_odds_manager
from yumachan import get_yumachan, get_race_course_from_cd
from ticket import make_ticket, make_csv
from util import get_time_stamp
from ipatgo import vote

TIME_STAMP = get_time_stamp()


def main(target_refund):
    start = time.time()

    # ゆまちゃんデータ取得
    yumachan = get_yumachan()

    # オッズを取得
    odds_manager = get_odds_manager(yumachan.opdt, yumachan.race_course, yumachan.rno)

    # 購入馬券リストを作る
    ticket_list = make_ticket(yumachan, odds_manager, target_refund)

    # 購入馬券リストをコンソール出力
    for ticket in ticket_list:
        print(ticket.to_string())

    # 購入馬券リストをcsvに書き出す
    make_csv(ticket_list, TIME_STAMP)

    # ipatgoで投票
    vote(TIME_STAMP)

    elapsed_time = time.time() - start
    print(f'完了！処理時間 : {round(elapsed_time, 2)}[秒]')


def only_odds_post(odds_manager, ws):
    print('オッズをエクセルに転記中...')
    horse_num = len(odds_manager.tan_odds_list)

    # 単勝オッズを転記
    for i, odds in enumerate(odds_manager.tan_odds_list):
        ws.cell(row=2 + i, column=4, value=float(odds.odds))

    # 複勝オッズを転記
    for i, odds in enumerate(odds_manager.fuku_min_odds_list):
        ws.cell(row=2 + i, column=16, value=float(odds.odds))

    # 馬連オッズを転記
    count = 0
    for i in range(horse_num - 1):
        for k in range(horse_num - i - 1):
            ws.cell(row=65 + k + i, column=2 + i, value=float(odds_manager.umaren_odds_list[count].odds))
            count += 1

    # ワイドオッズを転記
    count = 0
    for i in range(horse_num - 1):
        for k in range(horse_num - i - 1):
            ws.cell(row=145 + k + i, column=2 + i, value=float(odds_manager.wide_min_odds_list[count].odds))
            count += 1

    # 馬単オッズを転記
    count = 0
    for i in range(horse_num):
        for k in range(horse_num):
            if i == k:
                continue  # 軸紐が一致する箇所はスキップ
            ws.cell(row=225 + k, column=2 + i, value=float(odds_manager.umatan_odds_list[count].odds))
            count += 1

    # 3連複オッズを転記
    count = 0
    for i in range(horse_num):
        for k in range(horse_num - 1):
            for l in range(horse_num - k - 2):
                ws.cell(row=307 + 80 * i + k + l, column=2 + k,
                        value=float(odds_manager.trio_odds_list[count].odds))
                count += 1


def only_yuma_post(yumachan, ws):
    print('ゆまちゃんデータをエクセルに転記中...')
    # 馬の勝率を転記
    for horse in sorted(yumachan.horse_list):
        ws.cell(row=1 + int(horse.umano), column=9, value=float(horse.probability))


if __name__ == '__main__':
    if sys.argv[1] == 'odds_only':
        start = time.time()

        odds_manager = get_odds_manager(sys.argv[2], get_race_course_from_cd(sys.argv[3]), sys.argv[4])

        wb = openpyxl.load_workbook('xls/calc.xlsx')
        ws = wb['sheet']

        only_odds_post(odds_manager, ws)

        wb.save('xls/calc_odds.xlsx')
        wb.close()

        elapsed_time = time.time() - start
        print(f'完了！処理時間 : {round(elapsed_time, 2)}[秒]')
    elif sys.argv[1] == 'yuma_only':
        start = time.time()

        yumachan = get_yumachan()

        wb = openpyxl.load_workbook('xls/calc_odds.xlsx')
        ws = wb['sheet']

        only_yuma_post(yumachan, ws)

        wb.save('xls/calc_after.xlsx')
        wb.close()

        elapsed_time = time.time() - start
        print(f'完了！処理時間 : {round(elapsed_time, 2)}[秒]')
    elif sys.argv[1] == 'main':
        main(int(sys.argv[2]))

